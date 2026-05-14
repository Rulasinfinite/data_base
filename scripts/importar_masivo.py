"""
scripts/importar_masivo.py
SIDEC — Importación masiva de certificados (Excel)
Versión final: corrige extracción de IDENTIFICACIÓN.
"""

import os, sys, re, json, argparse
from datetime import datetime
from pathlib import Path
from concurrent.futures import ThreadPoolExecutor, as_completed

try:
    import psycopg2
    from psycopg2.extras import execute_values
except ImportError:
    print(json.dumps({"error": "psycopg2 no instalado"}), file=sys.stderr)
    sys.exit(1)

try:
    import openpyxl
except ImportError:
    print(json.dumps({"error": "openpyxl no instalado"}), file=sys.stderr)
    sys.exit(1)

import warnings
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

try:
    import xlrd
    TIENE_XLRD = True
except ImportError:
    TIENE_XLRD = False

try:
    from dateutil import parser
    TIENE_DATEUTIL = True
except ImportError:
    TIENE_DATEUTIL = False
    parser = None

DB_CONFIG = {
    'host':     os.getenv('DB_HOST',     'localhost'),
    'port':     int(os.getenv('DB_PORT', 5432)),
    'dbname':   os.getenv('DB_NAME',     'sidec_db'),
    'user':     os.getenv('DB_USER',     'postgres'),
    'password': os.getenv('DB_PASSWORD', 'sidecmexico'),
}

# ──────────── Helpers de extracción ────────────
def limpiar_valor(valor):
    """Limpia valores extraídos: elimina prefijos no deseados y espacios."""
    if valor is None:
        return None
    if isinstance(valor, datetime):
        return valor.strftime('%Y-%m-%d')
    val_str = str(valor).strip()
    if not val_str:
        return None
    # Eliminar prefijos comunes (incluyendo errores ortográficos)
    prefijos = [
        r'^del\s+instrumento:\s*',
        r'^descripción\s+del\s+instrumento:\s*',
        r'^instrument\s+description:\s*',
        r'^No:\s*',
        r'^Identificación\s+No:\s*',
        r'^identificacion\s+no:\s*',
        r'^entificación\s+No:\s*',          # <--- nuevo: error común
        r'^ID:\s*',
        r'^MAGNITUD\s+DE\s+EVALUADA:\s*',
        r'^magnitud\s+evaluada:\s*',
    ]
    for pref in prefijos:
        val_str = re.sub(pref, '', val_str, flags=re.IGNORECASE)
    # Si después de limpiar queda una cadena vacía o muy corta, descartar
    val_str = val_str.strip()
    if len(val_str) < 2:
        return None
    return val_str

# ── Patrón de rango de medición ──────────────────────────────────────────────
_UNIDADES = (
    r'kgf?|lbf?|g\b|mg\b|tonelada[s]?|t\b|N\b|kN\b|MN\b|'
    r'm\b|cm\b|mm\b|µm\b|nm\b|km\b|in\b|ft\b|yd\b|'
    r'°C|°F|K\b|'
    r'bar\b|mbar\b|psi\b|kPa\b|MPa\b|GPa\b|Pa\b|atm\b|'
    r'l\b|L\b|ml\b|mL\b|cl\b|'
    r'HRC\b|HRB\b|HRA\b|HV\b|HB\b|'
    r'Ω\b|kΩ\b|MΩ\b|V\b|mV\b|kV\b|A\b|mA\b|W\b|kW\b|Hz\b|kHz\b|MHz\b|'
    r'rpm\b|s\b|ms\b|min\b|h\b|%|ppm\b'
)
PATRON_RANGO = re.compile(
    r'\d+(?:[.,]\d+)?\s*(?:' + _UNIDADES + r')?\s*'
    r'(?:a\b|–|—|-|to\b)\s*'
    r'\d+(?:[.,]\d+)?\s*(?:' + _UNIDADES + r')?',
    re.IGNORECASE
)

def es_rango_valido(texto):
    """True si el texto contiene un patrón de rango numérico válido (ej. 0 kgf a 200 kgf)."""
    return bool(texto and PATRON_RANGO.search(str(texto)))

def extraer_rango_de_texto(texto):
    """Extrae el primer patrón de rango de medición del texto, o None si no lo hay."""
    if not texto:
        return None
    m = PATRON_RANGO.search(str(texto))
    return m.group(0).strip() if m else None

# ── Validación de Identificación ─────────────────────────────────────────────
_PALABRAS_DIRECCION = frozenset([
    'c.p.', ' cp ', 'código postal', 'colonia', 'col.', 'av.', 'ave.',
    'avenida', 'calle', 'carretera', 'blvd', 'boulevard', 'parque', 'plaza',
    'edificio', 'piso ', 'interior', 'local ', 'manzana', 'lote ',
    'fraccionamiento', 'ciudad de', 'estado de', 'municipio', 'delegación',
    'alcaldía', 'méxico', 'mexico', 'monterrey', 'guadalajara', 'saltillo',
    'industrial', 'zona ', 'sector ', 'km ', 'highway', 'street', 'avenue',
    'suite ', 'floor ', 'building', 'logística', 'logistica', 'parque ',
])
# Palabras que indican que se capturó la etiqueta en lugar del valor
_PALABRAS_ETIQUETA_ID = frozenset([
    'identificación', 'identificacion', 'identification',
    'entificación', 'entificacion',
    'id no', 'no.:', 'alcance', 'descripción', 'descripcion',
    'magnitud', 'modelo', 'marca', 'serie',
])

def es_identificacion_valida(valor):
    """
    Devuelve True si el valor parece un código de identificación de instrumento.
    Rechaza: textos >40 chars, multilinea, keywords de dirección o de etiqueta.
    """
    if not valor:
        return False
    s = str(valor).strip()
    if len(s) > 40:
        return False
    if '\n' in s or '\r' in s:
        return False
    s_lower = s.lower()
    for palabra in _PALABRAS_DIRECCION:
        if palabra in s_lower:
            return False
    for palabra in _PALABRAS_ETIQUETA_ID:
        if palabra in s_lower:
            return False
    return bool(re.search(r'[A-Za-z0-9]', s))

# ── Validación de Alcance de Medición ─────────────────────────────────────────
_PATRON_NUM_INFORME = re.compile(r'^[A-Za-z]{2,5}-\d+-\d+$')
_PATRON_SOLO_NUMEROS = re.compile(r'^\d+[\s\-]\d+$')
_PATRON_VALOR_CON_UNIDAD = re.compile(
    r'\d+(?:[.,]\d+)?\s*(?:' + _UNIDADES + r')', re.IGNORECASE
)

def es_alcance_valido(valor):
    """
    Acepta como alcance rangos numéricos ("0 kgf a 200 kgf") o valores
    simples con unidad ("18 lbf·ft", "100 kg").
    Rechaza patrones de número de informe ("LMM-026-18", "420-26") y texto puro.
    """
    if not valor:
        return False
    s = str(valor).strip()
    if _PATRON_NUM_INFORME.match(s):
        return False
    if _PATRON_SOLO_NUMEROS.match(s):
        return False
    if es_rango_valido(s):
        return True
    if _PATRON_VALOR_CON_UNIDAD.search(s):
        return True
    return False

# ── Campos faltantes → "N/A editar certificado" ───────────────────────────────
NA_EDITAR = "N/A editar certificado"
_CAMPOS_TEXTO = {
    'descripcion_instrumento', 'alcance', 'numero_serie', 'identificacion',
    'modelo', 'marca', 'magnitud_evaluada', 'resultado_calibracion',
    'incertidumbre', 'temperatura', 'humedad_relativa',
    'metodo_utilizado', 'lugar_calibracion', 'nombre_cliente',
}

def aplicar_na_faltantes(datos):
    """Rellena con NA_EDITAR los campos de texto que quedaron vacíos."""
    for campo in _CAMPOS_TEXTO:
        if not datos.get(campo):
            datos[campo] = NA_EDITAR
    return datos

def obtener_valor_celda(ws, row, col):
    try:
        return ws.cell(row=row, column=col).value
    except:
        return None

def extraer_desde_lineas(texto, etiqueta):
    """Busca la etiqueta en un texto multilínea y devuelve el valor asociado."""
    if not texto:
        return None
    lineas = texto.split('\n')
    for linea in lineas:
        limpia = linea.strip().lstrip('-•·*').strip()
        if etiqueta.lower() in limpia.lower():
            if ':' in limpia:
                _, val = limpia.split(':', 1)
                val = val.strip()
                if val:
                    return val
            else:
                idx = limpia.lower().find(etiqueta.lower())
                if idx != -1:
                    resto = limpia[idx + len(etiqueta):].strip()
                    if resto:
                        return resto
    return None

def normalizar_etiqueta(etiqueta):
    base = etiqueta.lower()
    variantes = {base}
    sin_acentos = base.replace('á','a').replace('é','e').replace('í','i').replace('ó','o').replace('ú','u')
    variantes.add(sin_acentos)
    variantes.add(base.replace(' ', ''))
    variantes.add(sin_acentos.replace(' ', ''))
    variantes.add(base.capitalize())
    variantes.add(base.upper())
    return variantes

def buscar_valor_en_fila(hoja, etiquetas, max_filas_extra=10, validador=None):
    """
    Busca una etiqueta (con variantes) y devuelve el valor asociado.
    Si se pasa `validador`, solo devuelve valores para los que validador(val) sea True.
    """
    todas_variantes = set()
    for eti in etiquetas:
        todas_variantes.update(normalizar_etiqueta(eti))

    def _aceptar(val):
        """Aplica limpiar_valor y, si hay validador, lo comprueba."""
        cleaned = limpiar_valor(val)
        if not cleaned:
            return None
        if validador is not None and not validador(cleaned):
            return None
        return cleaned

    ETIQUETAS_CONOCIDAS = {
        'nombre del cliente', 'customer name', 'client name', 'cliente',
        'dirección', 'direccion', 'address',
        'descripción del instrumento', 'instrument description', 'description',
        'alcance de medición', 'alcance de medicion', 'alcance', 'range', 'alcande',
        'no. de serie', 'serial number', 'serial no',
        'identificación no', 'identificacion no', 'identification',
        'modelo', 'model',
        'marca', 'manufacturer', 'brand',
        'magnitud evaluada', 'evaluated quality', 'magnitude',
        'resultado de calibración', 'calibration result',
        'incertidumbre', 'uncertainty',
        'temperatura', 'temperature',
        'humedad relativa', 'relative humidity', 'humidity',
        'fecha de emisión', 'fecha de emision', 'date of issue',
        'fecha de recepción', 'fecha de recepcion', 'reception date',
        'fecha de calibración', 'fecha de calibracion', 'date of calibration',
        'método utilizado', 'used method',
        'lugar de la calibración', 'place of calibration',
        'atención a', 'atencion a', 'attention to',
        'vigencia de calibración', 'vigency of calibration', 'vigencia'
    }

    def es_solo_etiqueta(texto):
        t = texto.strip().lower().rstrip(':').strip()
        return t in ETIQUETAS_CONOCIDAS or len(t) < 3

    for fila in hoja.iter_rows():
        for celda in fila:
            if celda.value and isinstance(celda.value, str):
                txt = str(celda.value)
                for variante in todas_variantes:
                    if variante in txt.lower():
                        # 0. Saltos de línea: para alcance, también buscar rango directo en el texto
                        if '\n' in txt:
                            val = extraer_desde_lineas(txt, variante)
                            if val:
                                r = _aceptar(val)
                                if r:
                                    return r
                            # Si el validador rechazó el resultado multilínea,
                            # intentar extraer un rango directamente de todo el texto
                            if validador is not None:
                                rango = extraer_rango_de_texto(txt)
                                if rango:
                                    r = _aceptar(rango)
                                    if r:
                                        return r
                        # 1. Después de ':'
                        if ':' in txt:
                            partes = txt.split(':', 1)
                            val = partes[1].strip() if len(partes) > 1 else None
                            if val and len(val) > 1:
                                r = _aceptar(val.split('\n')[0].strip())
                                if r:
                                    return r
                        # 2. Texto después de la etiqueta
                        idx = txt.lower().find(variante)
                        if idx != -1:
                            resto = txt[idx + len(variante):].strip()
                            if resto.startswith(':'):
                                resto = resto[1:].strip()
                            if resto and len(resto) > 1:
                                r = _aceptar(resto.split('\n')[0].strip())
                                if r:
                                    return r
                        # 3. Otra celda de la misma fila
                        for c in fila:
                            if c != celda and c.value:
                                val = limpiar_valor(c.value)
                                if val and len(val) > 1 and not es_solo_etiqueta(val):
                                    r = _aceptar(val)
                                    if r:
                                        return r
                        # 4. Celda siguiente
                        try:
                            sig = hoja.cell(row=celda.row, column=celda.column+1).value
                            if sig and not es_solo_etiqueta(str(sig)):
                                r = _aceptar(sig)
                                if r:
                                    return r
                        except:
                            pass
                        # 5. Filas siguientes
                        for offset in range(1, max_filas_extra+1):
                            try:
                                val = hoja.cell(row=celda.row+offset, column=celda.column).value
                                if val:
                                    val = limpiar_valor(val)
                                    if val and len(val) > 1 and not es_solo_etiqueta(val):
                                        r = _aceptar(val)
                                        if r:
                                            return r
                                val2 = hoja.cell(row=celda.row+offset, column=celda.column+1).value
                                if val2:
                                    val2 = limpiar_valor(val2)
                                    if val2 and len(val2) > 1 and not es_solo_etiqueta(val2):
                                        r = _aceptar(val2)
                                        if r:
                                            return r
                            except:
                                continue
    return None

def buscar_valor_multilenguaje(hoja, etiquetas_es, etiquetas_en=None):
    val = buscar_valor_en_fila(hoja, etiquetas_es)
    if val:
        return val, None
    if etiquetas_en:
        val = buscar_valor_en_fila(hoja, etiquetas_en)
        if val:
            return val, None
    return None, f"No se encontró: {', '.join(etiquetas_es + (etiquetas_en or []))}"

def limpiar_fecha(valor):
    """Convierte a date desde string, datetime o número Excel."""
    if not valor:
        return None
    if isinstance(valor, datetime):
        return valor.date()
    if isinstance(valor, (int, float)):
        try:
            from datetime import timedelta
            return datetime(1899, 12, 30) + timedelta(days=valor)
        except:
            pass
    valor_str = str(valor).strip()
    if not valor_str or len(valor_str) < 4:
        return None

    valor_str = re.sub(r'\s+de\s+', '-', valor_str, flags=re.IGNORECASE)
    valor_str = re.sub(r'\s+del?\s+', '-', valor_str, flags=re.IGNORECASE)

    meses = {
        'enero':1,'febrero':2,'marzo':3,'abril':4,'mayo':5,'junio':6,
        'julio':7,'agosto':8,'septiembre':9,'octubre':10,'noviembre':11,'diciembre':12,
        'ene':1,'feb':2,'mar':3,'abr':4,'may':5,'jun':6,'jul':7,'ago':8,'sep':9,'oct':10,'nov':11,'dic':12
    }
    m = re.match(r'([a-zA-Záéíóú]+)[\s-]+(\d{4})', valor_str, re.IGNORECASE)
    if m:
        mes_nombre = m.group(1).lower()
        año = int(m.group(2))
        if mes_nombre in meses:
            try:
                return datetime(año, meses[mes_nombre], 1).date()
            except:
                pass

    patron_fecha = re.compile(r'(\d{1,2}[\/\-\.]\d{1,2}[\/\-\.]\d{2,4}|\d{4}[\/\-\.]\d{1,2}[\/\-\.]\d{1,2})')
    match = patron_fecha.search(valor_str)
    if match:
        fecha_str = match.group(1)
        formatos = ['%Y-%m-%d','%Y/%m/%d','%d/%m/%Y','%d-%m-%Y','%d.%m.%Y','%m/%d/%Y']
        for fmt in formatos:
            try:
                return datetime.strptime(fecha_str, fmt).date()
            except:
                continue
    if TIENE_DATEUTIL and parser:
        try:
            if not re.search(r'\d', valor_str):
                return None
            fecha = parser.parse(valor_str, fuzzy=True, default=datetime(2000,1,1))
            if fecha.year < 1900 or fecha.year > 2100:
                return None
            return fecha.date()
        except:
            pass
    return None

def extraer_numero_informe(ws, ruta_archivo, identificacion_extraida=None):
    patron = r'(?:No\.?\s*)?([A-Za-z]{2,5}-\d+-\d+)'
    for fila in ws.iter_rows():
        for celda in fila:
            if celda.value and isinstance(celda.value, str):
                if 'hoja 1 de' in celda.value.lower():
                    for r in range(celda.row-1, celda.row+3):
                        for c in range(celda.column-2, celda.column+3):
                            try:
                                v = obtener_valor_celda(ws, r, c)
                                if v and isinstance(v, str):
                                    m = re.search(patron, v)
                                    if m:
                                        num = m.group(1)
                                        if not identificacion_extraida or num != identificacion_extraida:
                                            return num
                            except: continue
    for fila in ws.iter_rows():
        for celda in fila:
            if celda.value and isinstance(celda.value, str):
                m = re.search(patron, celda.value)
                if m:
                    num = m.group(1)
                    if not identificacion_extraida or num != identificacion_extraida:
                        return num
    nombre = Path(ruta_archivo).stem
    m = re.search(patron, nombre)
    if m:
        num = m.group(1)
        if not identificacion_extraida or num != identificacion_extraida:
            return num
    return nombre if nombre else None

def extraer_datos_de_hoja(ws, ruta_archivo):
    # --- Identificación mejorada ---
    # Lista ampliada de etiquetas posibles (incluyendo errores comunes)
    etiquetas_id = [
        "Identificación No", "Identificacion No", "IDENTIFICACIÓN", "ID", "Identification",
        "entificación No", "Identificación", "Identificacion", "ID No"
    ]
    identificacion, _ = buscar_valor_multilenguaje(ws, etiquetas_id, ["Identification", "ID"])
    # Validar que el resultado parezca realmente un código de instrumento
    if identificacion and not es_identificacion_valida(identificacion):
        identificacion = None
    if not identificacion:
        # Búsqueda más agresiva: buscar cualquier celda que contenga "identificación" y luego tomar el valor
        for fila in ws.iter_rows():
            for celda in fila:
                if celda.value and isinstance(celda.value, str):
                    txt = str(celda.value).lower()
                    if any(pat in txt for pat in ['identificación', 'identificacion', 'id']):
                        # Buscar después de ':' en la misma celda
                        if ':' in txt:
                            partes = txt.split(':', 1)
                            if len(partes) > 1:
                                cand = partes[1].strip()
                                if (cand and len(cand) > 1
                                        and not any(p in cand.lower() for p in ['identificación', 'identificacion'])
                                        and es_identificacion_valida(cand)):
                                    identificacion = cand
                                    break
                        # Buscar en celda derecha
                        sig = obtener_valor_celda(ws, celda.row, celda.column + 1)
                        if sig and isinstance(sig, str) and len(sig) > 1 and es_identificacion_valida(sig):
                            identificacion = sig
                            break
                        # Buscar abajo
                        abajo = obtener_valor_celda(ws, celda.row + 1, celda.column)
                        if abajo and isinstance(abajo, str) and len(abajo) > 1 and es_identificacion_valida(abajo):
                            identificacion = abajo
                            break
            if identificacion:
                break
    if identificacion:
        identificacion = limpiar_valor(identificacion)

    # Fechas
    fecha_emision = limpiar_fecha(buscar_valor_en_fila(ws, [
        "Fecha de emisión","Fecha de Emisión","Date of issue","FECHA DE EMISIÓN",
        "Fecha de emision","Fecha Emisión","Fecha Emision","Emisión","Emission date"
    ], max_filas_extra=10))
    if not fecha_emision:
        vigencia = buscar_valor_en_fila(ws, ["Vigencia de Calibración","Vigency of calibration","Vigencia"], max_filas_extra=5)
        fecha_emision = limpiar_fecha(vigencia)

    fecha_calibracion = limpiar_fecha(buscar_valor_en_fila(ws, [
        "Fecha de calibración","Fecha de Calibración","Date of calibration","FECHA DE CALIBRACIÓN",
        "Fecha de calibracion","Fecha Calibración","Fecha Calibracion","Fecha del estudio","Study date"
    ], max_filas_extra=10))

    fecha_recepcion = limpiar_fecha(buscar_valor_en_fila(ws, [
        "Fecha de recepción","Fecha de Recepción","Reception date","FECHA DE RECEPCIÓN",
        "Fecha de recepcion","Fecha Recepción","Fecha Recepcion"
    ], max_filas_extra=10))

    numero_informe = extraer_numero_informe(ws, ruta_archivo, identificacion)

    nombre_cliente, _ = buscar_valor_multilenguaje(ws,
        ["Nombre del cliente","Cliente","NOMBRE DEL CLIENTE","CLIENTE"],
        ["Customer name","Client name"])
    direccion = buscar_valor_en_fila(ws, ["Dirección","Direccion","Address","DIRECCIÓN"], max_filas_extra=4)
    atencion_a = buscar_valor_en_fila(ws, ["Atención a","Atencion a","Attention to","ATENCIÓN A"])

    # Descripción del instrumento (ya funciona)
    descripcion, _ = buscar_valor_multilenguaje(ws,
        ["Descripción del instrumento","Descripción","DESCRIPCIÓN","DESCRIPCION","Instrument description","Description"],
        [])
    if not descripcion or descripcion.lower() in ('del instrumento', 'del instrumento:', 'instrument description', 'descripción', 'descripción del instrumento'):
        descripcion = None
        for fila in ws.iter_rows():
            for celda in fila:
                if celda.value and isinstance(celda.value, str):
                    txt = str(celda.value).lower()
                    if any(pat in txt for pat in ['descripción del instrumento', 'descripción', 'instrument description']):
                        if ':' in txt:
                            partes = txt.split(':', 1)
                            if len(partes) > 1:
                                cand = partes[1].strip()
                                if cand and len(cand) > 2 and 'descripción' not in cand.lower():
                                    descripcion = cand
                                    break
                        for offset in range(1, 5):
                            val_derecha = obtener_valor_celda(ws, celda.row, celda.column + offset)
                            if val_derecha and isinstance(val_derecha, str) and len(val_derecha) > 2:
                                if not any(p in val_derecha.lower() for p in ['descripción', 'instrument']):
                                    descripcion = val_derecha
                                    break
                            val_abajo = obtener_valor_celda(ws, celda.row + offset, celda.column)
                            if val_abajo and isinstance(val_abajo, str) and len(val_abajo) > 2:
                                if not any(p in val_abajo.lower() for p in ['descripción', 'instrument']):
                                    descripcion = val_abajo
                                    break
                        if descripcion:
                            break
            if descripcion:
                break
    if descripcion:
        descripcion = limpiar_valor(descripcion)

    numero_serie, _ = buscar_valor_multilenguaje(ws,
        ["No. De Serie","No. de serie","NO. DE SERIE","NÚMERO DE SERIE","Serial number","Serial No"],
        [])
    modelo, _ = buscar_valor_multilenguaje(ws, ["Modelo","MODELO"], ["Model"])
    marca, _ = buscar_valor_multilenguaje(ws, ["Marca","MARCA"], ["Manufacturer","Brand"])
    magnitud, _ = buscar_valor_multilenguaje(ws,
        ["Magnitud evaluada","MAGNITUD EVALUADA","Evaluated quality","Magnitude"],
        [])
    resultado, _ = buscar_valor_multilenguaje(ws,
        ["Resultado de calibración","Resultado del estudio","Calibration result","Study result"],
        [])
    lugar, _ = buscar_valor_multilenguaje(ws,
        ["Lugar de la calibración","Lugar del estudio","Place of calibration","Study place"],
        [])

    # Alcance — buscar con validación de patrón de rango numérico o valor con unidad
    alcance = buscar_valor_en_fila(ws, [
        "Alcance de Medición","Alcance de medicion","ALCANCE DE MEDICIÓN",
        "Alcance de Medicion","Range of measurement","Range","Alcande","Rango",
        "Alcance"
    ], max_filas_extra=8, validador=es_alcance_valido)

    # Fallback 1: buscar celdas con etiqueta de alcance y extraer rango del texto completo
    if not alcance:
        _etq_alcance = ['alcance de medición', 'alcance de medicion', 'alcance', 'range', 'alcande', 'rango']
        for fila in ws.iter_rows():
            for celda in fila:
                if celda.value and isinstance(celda.value, str):
                    txt = celda.value
                    if any(e in txt.lower() for e in _etq_alcance):
                        rng = extraer_rango_de_texto(txt)
                        if rng and es_alcance_valido(rng):
                            alcance = rng
                            break
                        # También intentar leer valor con unidad de la misma celda
                        m_unidad = _PATRON_VALOR_CON_UNIDAD.search(txt)
                        if m_unidad and es_alcance_valido(m_unidad.group(0)):
                            alcance = m_unidad.group(0).strip()
                            break
            if alcance:
                break

    # Fallback 2: buscar en toda la hoja cualquier celda que contenga un patrón de rango
    if not alcance:
        for fila in ws.iter_rows():
            for celda in fila:
                if celda.value and isinstance(celda.value, str):
                    rng = extraer_rango_de_texto(celda.value)
                    if rng and es_alcance_valido(rng):
                        alcance = rng
                        break
            if alcance:
                break

    # Temperatura
    temperatura = buscar_valor_en_fila(ws, ["Temperatura","TEMPERATURA","Temperature"], max_filas_extra=8)
    if not temperatura:
        for fila in ws.iter_rows():
            for celda in fila:
                if celda.value and isinstance(celda.value, str):
                    if any(t in celda.value.lower() for t in ['temperatura','temperature']):
                        for r in range(celda.row, celda.row+5):
                            for c in range(celda.column-1, celda.column+5):
                                try:
                                    v = obtener_valor_celda(ws, r, c)
                                    if v and isinstance(v, str) and '°C' in v:
                                        temperatura = v.strip()
                                        break
                                except: pass
                            if temperatura: break
                        if not temperatura:
                            for c in fila:
                                if c != celda and c.value and isinstance(c.value, str):
                                    if '°C' in c.value:
                                        temperatura = c.value.strip()
                                        break
                    if temperatura: break
            if temperatura: break

    # Humedad
    humedad = buscar_valor_en_fila(ws, [
        "Humedad Relativa","HUMEDAD RELATIVA","Humedad","Relative Humidity","Humidity"
    ], max_filas_extra=8)
    if not humedad:
        for fila in ws.iter_rows():
            for celda in fila:
                if celda.value and isinstance(celda.value, str):
                    if any(h in celda.value.lower() for h in ['humedad','humidity']):
                        for r in range(celda.row, celda.row+5):
                            for c in range(celda.column-1, celda.column+5):
                                try:
                                    v = obtener_valor_celda(ws, r, c)
                                    if v and isinstance(v, str) and '%' in v:
                                        humedad = v.strip()
                                        break
                                except: pass
                            if humedad: break
                        if not humedad:
                            for c in fila:
                                if c != celda and c.value and isinstance(c.value, str) and '%' in c.value:
                                    humedad = c.value.strip()
                                    break
                    if humedad: break
            if humedad: break

    metodo = (buscar_valor_en_fila(ws, ["Método utilizado","Método Utilizado","Used method","Procedimiento"], max_filas_extra=5) or
              buscar_valor_multilenguaje(ws, ["Método utilizado"], ["Used method"])[0])
    incertidumbre = (buscar_valor_en_fila(ws, [
        "Incertidumbre","Uncertainty","INCERTIDUMBRE",
        "Incertidumbre de medición","Measurement uncertainty"
    ], max_filas_extra=8) or
                     buscar_valor_multilenguaje(ws, ["Incertidumbre"], ["Uncertainty"])[0])
    # Si la hoja delega la incertidumbre a otra hoja, conservar ese texto
    if not incertidumbre:
        for fila in ws.iter_rows():
            for celda in fila:
                if celda.value and isinstance(celda.value, str):
                    txt = celda.value
                    if 'incertidumbre' in txt.lower() or 'uncertainty' in txt.lower():
                        if 'hoja' in txt.lower() or 'sheet' in txt.lower() or 'indica' in txt.lower():
                            incertidumbre = txt.strip()
                            break
            if incertidumbre:
                break

    anio = None
    if fecha_emision:
        anio = fecha_emision.year
    elif numero_informe:
        m = re.search(r'-(\d{2})$', numero_informe)
        if m:
            anio = 2000 + int(m.group(1)) if int(m.group(1)) <= 50 else 1900 + int(m.group(1))
    if not anio:
        anio = datetime.now().year

    datos = {
        "numero_informe": numero_informe,
        "anio_emision": anio,
        "nombre_cliente": nombre_cliente,
        "direccion": direccion,
        "atencion_a": atencion_a,
        "descripcion_instrumento": descripcion,
        "alcance": alcance,
        "numero_serie": numero_serie,
        "identificacion": identificacion,
        "modelo": modelo,
        "marca": marca,
        "magnitud_evaluada": magnitud,
        "resultado_calibracion": resultado,
        "incertidumbre": incertidumbre,
        "temperatura": temperatura,
        "humedad_relativa": humedad,
        "metodo_utilizado": metodo,
        "lugar_calibracion": lugar,
        "calibrado_por": None,
        "aprobado_por": None,
        "fecha_recepcion": fecha_recepcion,
        "fecha_calibracion": fecha_calibracion,
        "fecha_emision": fecha_emision,
    }
    return aplicar_na_faltantes(datos)

def extraer_datos_excel(ruta_archivo):
    ext = Path(ruta_archivo).suffix.lower()
    if ext == '.xls' and not ruta_archivo.lower().endswith('.xlsx'):
        return extraer_datos_xls(ruta_archivo)

    try:
        wb = openpyxl.load_workbook(ruta_archivo, data_only=True)
        for nombre in wb.sheetnames:
            if any(p in nombre.lower() for p in ['portada', 'informe de calibración', 'certificado']):
                ws = wb[nombre]
                datos = extraer_datos_de_hoja(ws, ruta_archivo)
                if datos:
                    datos["ruta_archivo_origen"] = ruta_archivo
                    datos["fecha_importacion"] = datetime.now()
                    return datos, None
        for nombre in wb.sheetnames:
            ws = wb[nombre]
            datos = extraer_datos_de_hoja(ws, ruta_archivo)
            if datos:
                datos["ruta_archivo_origen"] = ruta_archivo
                datos["fecha_importacion"] = datetime.now()
                return datos, None
        return None, "No se encontró hoja de portada ni datos en ninguna hoja."
    except Exception as e:
        return None, f"Error Excel: {str(e)}"

# ═══════════════ Funciones para .xls antiguo (con corrección de identificación) ═══════════════
def extraer_datos_xls(ruta_archivo):
    if not TIENE_XLRD:
        return None, "xlrd no instalado."
    try:
        wb = xlrd.open_workbook(ruta_archivo)
        for n in wb.sheet_names():
            ws = wb.sheet_by_name(n)
            datos = _extraer_xls_desde_ws(ws, ruta_archivo)
            if datos[0] is not None:
                return datos
        return None, "No se encontraron datos en ninguna hoja"
    except Exception as e:
        return None, f"Error XLS: {str(e)}"

def _extraer_xls_desde_ws(ws, ruta_archivo):
    nrows = ws.nrows
    ncols = ws.ncols
    datos_crudos = []
    for r in range(nrows):
        fila = []
        for c in range(ncols):
            v = ws.cell_value(r,c)
            if isinstance(v, float) and v == int(v):
                v = str(int(v))
            elif isinstance(v, float):
                v = str(v)
            else:
                v = str(v) if v else None
            fila.append(v)
        datos_crudos.append(fila)

    def _get_val(row, col):
        if 0 <= row < nrows and 0 <= col < ncols:
            return datos_crudos[row][col]
        return None

    def _es_etiqueta_pura(texto):
        if not texto: return False
        t = texto.strip().lower().rstrip(':').strip()
        etiq = {
            'nombre del cliente', 'cliente', 'dirección', 'direccion', 'address',
            'descripción del instrumento', 'instrument description', 'alcance de medición',
            'alcance', 'range', 'no. de serie', 'serial number', 'identificación no',
            'identificacion no', 'modelo', 'marca', 'manufacturer', 'magnitud evaluada',
            'resultado de calibración', 'incertidumbre', 'temperatura', 'temperature',
            'humedad relativa', 'fecha de emisión', 'fecha de recepción', 'fecha de calibración',
            'método utilizado', 'lugar de la calibración', 'atención a', 'vigencia'
        }
        return t in etiq or len(t) < 3

    def buscar_valor_xls(etiquetas, max_filas_extra=10, validador=None):
        todas_variantes = set()
        for eti in etiquetas:
            todas_variantes.update(normalizar_etiqueta(eti))

        def _aceptar_xls(val):
            cleaned = limpiar_valor(val)
            if not cleaned:
                return None
            if validador is not None and not validador(cleaned):
                return None
            return cleaned

        for r in range(nrows):
            for c in range(ncols):
                txt = _get_val(r,c)
                if not txt or not isinstance(txt, str):
                    continue
                for variante in todas_variantes:
                    if variante in txt.lower():
                        if '\n' in txt:
                            val = extraer_desde_lineas(txt, variante)
                            if val:
                                result = _aceptar_xls(val)
                                if result:
                                    return result
                            if validador is not None:
                                rng = extraer_rango_de_texto(txt)
                                if rng:
                                    result = _aceptar_xls(rng)
                                    if result:
                                        return result
                        if ':' in txt:
                            partes = txt.split(':', 1)
                            if len(partes) > 1:
                                val = partes[1].strip()
                                if val and len(val) > 1:
                                    result = _aceptar_xls(val.split('\n')[0].strip())
                                    if result:
                                        return result
                        idx = txt.lower().find(variante)
                        if idx != -1:
                            resto = txt[idx+len(variante):].strip()
                            if resto.startswith(':'):
                                resto = resto[1:].strip()
                            if resto and len(resto) > 1:
                                result = _aceptar_xls(resto.split('\n')[0].strip())
                                if result:
                                    return result
                        for cc in range(ncols):
                            if cc != c:
                                val = _get_val(r, cc)
                                if val and not _es_etiqueta_pura(val):
                                    result = _aceptar_xls(val)
                                    if result:
                                        return result
                        val_sig = _get_val(r, c+1)
                        if val_sig and not _es_etiqueta_pura(val_sig):
                            result = _aceptar_xls(val_sig)
                            if result:
                                return result
                        for offset in range(1, max_filas_extra+1):
                            val_bajo = _get_val(r+offset, c)
                            if val_bajo and not _es_etiqueta_pura(val_bajo):
                                result = _aceptar_xls(val_bajo)
                                if result:
                                    return result
                            val_bajo2 = _get_val(r+offset, c+1)
                            if val_bajo2 and not _es_etiqueta_pura(val_bajo2):
                                result = _aceptar_xls(val_bajo2)
                                if result:
                                    return result
        return None

    def multileng_xls(es, en=None):
        v = buscar_valor_xls(es)
        if v: return v, None
        if en:
            v = buscar_valor_xls(en)
            if v: return v, None
        return None, None

    # --- Identificación mejorada para .xls ---
    etiquetas_id_xls = [
        "Identificación No", "Identificacion No", "IDENTIFICACIÓN", "ID", "Identification",
        "entificación No", "Identificación", "Identificacion"
    ]
    identificacion, _ = multileng_xls(etiquetas_id_xls, ["Identification", "ID"])
    # Validar que el resultado no sea una dirección u otro texto largo
    if identificacion and not es_identificacion_valida(identificacion):
        identificacion = None
    if not identificacion:
        for r in range(nrows):
            for c in range(ncols):
                txt = _get_val(r,c)
                if txt and isinstance(txt, str):
                    if any(pat in txt.lower() for pat in ['identificación', 'identificacion', 'id']):
                        if ':' in txt:
                            partes = txt.split(':', 1)
                            if len(partes) > 1:
                                cand = partes[1].strip()
                                if cand and len(cand) > 1 and es_identificacion_valida(cand):
                                    identificacion = cand
                                    break
                        sig = _get_val(r, c+1)
                        if sig and isinstance(sig, str) and len(sig) > 1 and es_identificacion_valida(sig):
                            identificacion = sig
                            break
                        abajo = _get_val(r+1, c)
                        if abajo and isinstance(abajo, str) and len(abajo) > 1 and es_identificacion_valida(abajo):
                            identificacion = abajo
                            break
            if identificacion:
                break
    if identificacion:
        identificacion = limpiar_valor(identificacion)

    nombre_cliente, _ = multileng_xls(["Nombre del cliente","Cliente","NOMBRE DEL CLIENTE"], ["Customer name","Client name"])
    descripcion, _ = multileng_xls(["Descripción del instrumento","Descripción","DESCRIPCIÓN","Instrument description"], [])
    if descripcion:
        descripcion = limpiar_valor(descripcion)
    else:
        # búsqueda adicional para descripción (igual que antes)
        for r in range(nrows):
            for c in range(ncols):
                txt = _get_val(r,c)
                if txt and isinstance(txt, str):
                    if any(p in txt.lower() for p in ['descripción del instrumento', 'instrument description']):
                        if ':' in txt:
                            partes = txt.split(':', 1)
                            if len(partes) > 1:
                                cand = partes[1].strip()
                                if cand and 'descripción' not in cand.lower():
                                    descripcion = limpiar_valor(cand)
                                    break
                        for offset in range(1, 5):
                            vd = _get_val(r, c+offset)
                            if vd and len(vd) > 2 and 'descripción' not in vd.lower():
                                descripcion = limpiar_valor(vd)
                                break
                            va = _get_val(r+offset, c)
                            if va and len(va) > 2 and 'descripción' not in va.lower():
                                descripcion = limpiar_valor(va)
                                break
            if descripcion:
                break

    direccion = buscar_valor_xls(["Dirección","Direccion","Address","DIRECCIÓN"])
    atencion_a = buscar_valor_xls(["Atención a","Atencion a","Attention to"])

    fecha_em = limpiar_fecha(buscar_valor_xls([
        "Fecha de emisión","Fecha de Emisión","Date of issue",
        "Fecha de emision","Fecha Emisión","Fecha Emision",
        "FECHA DE EMISIÓN","Emisión","Vigencia de Calibración","Vigencia"
    ], max_filas_extra=10))

    fecha_rec = limpiar_fecha(buscar_valor_xls([
        "Fecha de recepción","Fecha de Recepción","Reception date",
        "Fecha de recepcion","Fecha Recepción"
    ], max_filas_extra=10))

    fecha_cal = limpiar_fecha(buscar_valor_xls([
        "Fecha de calibración","Fecha de Calibración","Date of calibration",
        "Fecha de calibracion","Fecha Calibración","Fecha del estudio"
    ], max_filas_extra=10))

    num_info = None
    patron_num = r'([A-Za-z]{2,5}-\d+-\d+)'
    for r in range(nrows):
        for c in range(ncols):
            txt = _get_val(r,c)
            if txt:
                m = re.search(patron_num, str(txt))
                if m:
                    cand = m.group(1)
                    if not identificacion or cand != identificacion:
                        num_info = cand
                        break
        if num_info: break
    if not num_info:
        nombre = Path(ruta_archivo).stem
        m = re.search(patron_num, nombre)
        num_info = m.group(1) if m else nombre

    anio = None
    if fecha_em: anio = fecha_em.year
    elif num_info:
        m = re.search(r'-(\d{2})$', num_info)
        if m: anio = 2000+int(m.group(1)) if int(m.group(1))<=50 else 1900+int(m.group(1))
    if not anio: anio = datetime.now().year

    # Alcance con validador unificado (rango o valor-con-unidad)
    alcance = buscar_valor_xls([
        "Alcance de Medición","Alcance de medicion","ALCANCE DE MEDICIÓN",
        "Range of measurement","Range","Alcande","Rango","Alcance"
    ], max_filas_extra=8, validador=es_alcance_valido)

    # Fallback 1: celdas con etiqueta de alcance — extraer rango o valor con unidad
    if not alcance:
        _etq_alcance = ['alcance de medición', 'alcance de medicion', 'alcance', 'range', 'alcande', 'rango']
        for r in range(nrows):
            for c in range(ncols):
                txt = _get_val(r, c)
                if txt and isinstance(txt, str) and any(e in txt.lower() for e in _etq_alcance):
                    rng = extraer_rango_de_texto(txt)
                    if rng and es_alcance_valido(rng):
                        alcance = rng
                        break
                    m_unidad = _PATRON_VALOR_CON_UNIDAD.search(txt)
                    if m_unidad and es_alcance_valido(m_unidad.group(0)):
                        alcance = m_unidad.group(0).strip()
                        break
            if alcance:
                break

    # Fallback 2: buscar en toda la hoja cualquier celda con patrón de rango numérico
    if not alcance:
        for r in range(nrows):
            for c in range(ncols):
                txt = _get_val(r, c)
                if txt and isinstance(txt, str):
                    rng = extraer_rango_de_texto(txt)
                    if rng and es_alcance_valido(rng):
                        alcance = rng
                        break
            if alcance:
                break

    numero_serie, _ = multileng_xls(["No. De Serie","No. de serie","SERIAL","Serial number"], [])
    modelo, _ = multileng_xls(["Modelo"], ["Model"])
    marca, _ = multileng_xls(["Marca"], ["Manufacturer","Brand"])
    magnitud, _ = multileng_xls(["Magnitud evaluada","MAGNITUD","Evaluated quality"], [])
    resultado, _ = multileng_xls(["Resultado de calibración","Resultado del estudio","Calibration result"], [])
    lugar, _ = multileng_xls(["Lugar de la calibración","Lugar del estudio","Place of calibration"], [])
    metodo, _ = multileng_xls(["Método utilizado","MÉTODO","Used method","Procedimiento"], [])
    incertidumbre, _ = multileng_xls([
        "Incertidumbre","INCERTIDUMBRE","Uncertainty",
        "Incertidumbre de medición","Measurement uncertainty"
    ], [])
    # Si la hoja delega la incertidumbre a otra hoja, conservar ese texto
    if not incertidumbre:
        for r in range(nrows):
            for c in range(ncols):
                txt = _get_val(r, c)
                if txt and isinstance(txt, str):
                    if ('incertidumbre' in txt.lower() or 'uncertainty' in txt.lower()):
                        if ('hoja' in txt.lower() or 'sheet' in txt.lower() or 'indica' in txt.lower()):
                            incertidumbre = txt.strip()
                            break
            if incertidumbre:
                break

    temperatura = buscar_valor_xls(["Temperatura","TEMPERATURA","Temperature"], max_filas_extra=8)
    humedad = buscar_valor_xls(["Humedad Relativa","Humedad","Relative Humidity","Humidity"], max_filas_extra=8)

    datos = {
        "numero_informe": num_info,
        "anio_emision": anio,
        "nombre_cliente": nombre_cliente,
        "direccion": direccion,
        "atencion_a": atencion_a,
        "descripcion_instrumento": descripcion,
        "alcance": alcance,
        "numero_serie": numero_serie,
        "identificacion": identificacion,
        "modelo": modelo,
        "marca": marca,
        "magnitud_evaluada": magnitud,
        "resultado_calibracion": resultado,
        "incertidumbre": incertidumbre,
        "temperatura": temperatura,
        "humedad_relativa": humedad,
        "metodo_utilizado": metodo,
        "lugar_calibracion": lugar,
        "calibrado_por": None,
        "aprobado_por": None,
        "fecha_recepcion": fecha_rec,
        "fecha_calibracion": fecha_cal,
        "fecha_emision": fecha_em,
        "ruta_archivo_origen": ruta_archivo,
        "fecha_importacion": datetime.now(),
    }
    aplicar_na_faltantes(datos)
    return datos, None

# ═══════════════ Validación e Inserción (sin cambios) ═══════════════
CAMPOS_PRINCIPALES = [
    'numero_informe', 'nombre_cliente', 'descripcion_instrumento',
    'alcance', 'numero_serie', 'identificacion', 'modelo', 'marca',
    'magnitud_evaluada', 'fecha_emision', 'fecha_calibracion',
    'lugar_calibracion'
]

def es_certificado_valido(datos):
    presentes = sum(1 for campo in CAMPOS_PRINCIPALES if datos.get(campo))
    return presentes >= 8

def escanear_carpeta(carpeta, anio_filtro=None):
    archivos = []
    raiz = Path(carpeta)
    if not raiz.exists(): return []
    for ruta in raiz.rglob('*'):
        if not ruta.is_file() or ruta.name.startswith('~$'): continue
        ext = ruta.suffix.lower()
        if ext not in ('.xlsx','.xls'): continue
        if anio_filtro:
            m = re.search(r'20[1-3]\d', str(ruta))
            if m and int(m.group())!=int(anio_filtro): continue
        archivos.append(str(ruta))
    return sorted(archivos)

def resolver_clientes_en_lote(cursor, datos_lista):
    clientes = {}
    for d in datos_lista:
        nombre = d.get('nombre_cliente') or 'Cliente no identificado'
        direccion = d.get('direccion')
        atencion = d.get('atencion_a')
        if nombre not in clientes:
            clientes[nombre] = {'nombre': nombre, 'direccion': direccion, 'atencion': atencion}
        else:
            if not clientes[nombre]['direccion'] and direccion:
                clientes[nombre]['direccion'] = direccion
            if not clientes[nombre]['atencion'] and atencion:
                clientes[nombre]['atencion'] = atencion

    nombres_unicos = list(clientes.keys())
    cursor.execute("SELECT nombre, id FROM clientes WHERE nombre = ANY(%s)", (nombres_unicos,))
    existentes = {row[0]: row[1] for row in cursor.fetchall()}

    nuevos_nombres = [n for n in nombres_unicos if n not in existentes]
    if nuevos_nombres:
        inserts = [(clientes[n]['nombre'], clientes[n]['direccion'], clientes[n]['atencion']) for n in nuevos_nombres]
        execute_values(cursor,
            "INSERT INTO clientes (nombre, direccion, atencion_a) VALUES %s ON CONFLICT (nombre) DO UPDATE SET direccion = EXCLUDED.direccion, atencion_a = EXCLUDED.atencion_a",
            inserts
        )
        cursor.execute("SELECT nombre, id FROM clientes WHERE nombre = ANY(%s)", (nuevos_nombres,))
        existentes.update({row[0]: row[1] for row in cursor.fetchall()})

    for d in datos_lista:
        nombre = d.get('nombre_cliente') or 'Cliente no identificado'
        d['cliente_id'] = existentes[nombre]
        if not d.get('numero_informe'):
            d['numero_informe'] = Path(d.get('ruta_archivo_origen', '')).stem

def insertar_certificados(conn, datos_lista, empleado=None):
    if not datos_lista:
        return 0, 0, [], []

    cursor = conn.cursor()
    resolver_clientes_en_lote(cursor, datos_lista)

    if empleado:
        for d in datos_lista:
            d['importado_por'] = empleado

    columnas = [
        'numero_informe','anio_emision','cliente_id',
        'descripcion_instrumento','alcance','numero_serie',
        'identificacion','modelo','marca','magnitud_evaluada',
        'resultado_calibracion','incertidumbre','temperatura','humedad_relativa',
        'metodo_utilizado','lugar_calibracion',
        'calibrado_por','aprobado_por',
        'fecha_recepcion','fecha_calibracion','fecha_emision',
        'estado','fecha_vencimiento',
        'ruta_archivo_origen','fecha_importacion','importado_por'
    ]

    combinaciones = [(d['numero_informe'], d['cliente_id'], d.get('marca')) for d in datos_lista]
    cursor.execute(
        "SELECT numero_informe, cliente_id, marca FROM certificados WHERE (numero_informe, cliente_id, marca) IN %s",
        (tuple(combinaciones),)
    )
    existentes = set((row[0], row[1], row[2]) for row in cursor.fetchall())

    nuevos = []
    duplicados = []
    for d in datos_lista:
        clave = (d['numero_informe'], d['cliente_id'], d.get('marca'))
        if clave in existentes:
            duplicados.append(d)
        else:
            nuevos.append(d)

    errores = []
    if nuevos:
        try:
            valores = [tuple(d.get(c) for c in columnas) for d in nuevos]
            execute_values(cursor,
                f"INSERT INTO certificados ({', '.join(columnas)}) VALUES %s "
                f"ON CONFLICT (numero_informe, cliente_id, marca) DO NOTHING",
                valores
            )
            conn.commit()
        except Exception as e:
            conn.rollback()
            errores.append({"archivo": "lote", "error": str(e)})
            print(f"Error en inserción masiva: {e}", file=sys.stderr)

    cursor.close()
    exitosos = len(nuevos)
    omitidos = len(duplicados)
    archivos_omitidos = [Path(d.get('ruta_archivo_origen', '')).name for d in duplicados]
    return exitosos, omitidos, archivos_omitidos, errores

# ── MAIN ──
if __name__ == '__main__':
    parser = argparse.ArgumentParser()
    parser.add_argument('--json-output', action='store_true')
    parser.add_argument('--carpeta', required=True)
    parser.add_argument('--empleado', default=None)
    parser.add_argument('--anio', type=int, default=None)
    parser.add_argument('--solo-escanear', action='store_true')
    args = parser.parse_args()

    archivos = escanear_carpeta(args.carpeta, args.anio)

    if args.solo_escanear:
        por_tipo = {'excel':0}
        for a in archivos:
            ext = Path(a).suffix.lower()
            if ext in ('.xlsx','.xls'): por_tipo['excel']+=1
        res = {"total":len(archivos),"por_tipo":por_tipo,"archivos":[Path(a).name for a in archivos[:20]]}
        if args.json_output:
            print('__JSON_START__'); print(json.dumps(res,default=str)); print('__JSON_END__')
        else: print(json.dumps(res,default=str))
        sys.exit(0)

    try:
        conn = psycopg2.connect(**DB_CONFIG)
    except Exception as e:
        print(json.dumps({"error":f"No se pudo conectar a PostgreSQL: {e}"}))
        sys.exit(1)

    datos_validos = []
    archivos_exitosos = []
    archivos_no_validos = []
    errores_extraccion = []
    no_certs = 0
    total = len(archivos)

    max_workers = min(4, total) if total > 0 else 1
    completados = 0

    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        futures = {executor.submit(extraer_datos_excel, archivo): archivo for archivo in archivos}

        for future in as_completed(futures):
            archivo = futures[future]
            completados += 1

            try:
                datos, error = future.result()
            except Exception as e:
                datos = None
                error = str(e)

            print(f"Procesando: {archivo}", file=sys.stderr, flush=True)

            if error:
                errores_extraccion.append({"archivo": archivo, "error": error})
            elif datos:
                if not datos.get('numero_informe'):
                    datos['numero_informe'] = os.path.basename(archivo).rsplit('.', 1)[0]
                if es_certificado_valido(datos):
                    datos_validos.append(datos)
                    archivos_exitosos.append(os.path.basename(archivo))
                else:
                    archivos_no_validos.append(os.path.basename(archivo))
                    no_certs += 1
            else:
                archivos_no_validos.append(os.path.basename(archivo))
                no_certs += 1

            if completados % 5 == 0 or completados == 1 or completados == total:
                print(f"Procesados {completados}/{total} archivos...", file=sys.stderr, flush=True)

    exitosos, omitidos, archivos_omitidos, errores_insercion = insertar_certificados(conn, datos_validos, args.empleado)
    conn.close()

    todos_errores = errores_extraccion + errores_insercion

    res = {
        "total": total,
        "exitosos": exitosos,
        "omitidos": omitidos,
        "fallidos": len(todos_errores),
        "no_certificados": no_certs,
        "archivos_exitosos": archivos_exitosos,
        "archivos_omitidos": archivos_omitidos,
        "archivos_no_validos": archivos_no_validos,
        "errores": todos_errores[:50] if todos_errores else []
    }
    if args.json_output:
        print('__JSON_START__')
        print(json.dumps(res, default=str))
        print('__JSON_END__')
    else:
        print(json.dumps(res, default=str))