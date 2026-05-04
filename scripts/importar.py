"""
scripts/importar.py
SIDEC — Importación masiva de certificados de calibración
Basado en el script original, adaptado para PostgreSQL + manejo robusto de 40+ GB

Uso:
    pip install pandas openpyxl psycopg2-binary tqdm
    python importar.py

    O con argumentos:
    python importar.py --carpeta "C:/Instrumentos/2017" --anio 2017
"""

import os
import sys
import argparse
import pandas as pd
import psycopg2
from psycopg2.extras import execute_values
from openpyxl import load_workbook
from datetime import datetime
from tqdm import tqdm
import json
import logging

# ============================================================
# Configuración
# ============================================================
DB_CONFIG = {
    'host':     os.getenv('DB_HOST',     'sidecmexico'),
    'port':     int(os.getenv('DB_PORT', 5432)),
    'dbname':   os.getenv('DB_NAME',     'sidec_db'),
    'user':     os.getenv('DB_USER',     'postgres'),
    'password': os.getenv('DB_PASSWORD', ''),
}

# Carpetas a procesar (igual que el script original)
CARPETAS_DEFAULT = [
    r'C:\PruebaLocal',
    # Descomenta según necesites:
    # r'\\PC1\Instrumentos',
    # r'\\PC2\Instrumentos',
    # r'\\PC3\Instrumentos',
    # r'\\PC4\Instrumentos',
    # r'\\PC5\Instrumentos',
]

CHUNK_SIZE = 500  # Insertar en lotes de 500 para no saturar la DB

# ============================================================
# Logger
# ============================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s [%(levelname)s] %(message)s',
    handlers=[
        logging.FileHandler(f'importacion_{datetime.now().strftime("%Y%m%d_%H%M%S")}.log'),
        logging.StreamHandler(sys.stdout),
    ]
)
log = logging.getLogger(__name__)

# ============================================================
# Funciones de extracción (misma lógica del script original)
# ============================================================
def buscar_valor_en_fila(hoja, texto_clave):
    """Busca una celda con el texto clave y devuelve el valor de la siguiente celda en la misma fila."""
    for fila in hoja.iter_rows():
        celda_clave = None
        for celda in fila:
            try:
                if celda.value and isinstance(celda.value, str) and texto_clave.lower() in celda.value.lower():
                    celda_clave = celda
                    break
            except:
                continue
        if celda_clave:
            for celda_misma_fila in fila:
                try:
                    if celda_misma_fila != celda_clave and celda_misma_fila.value:
                        return str(celda_misma_fila.value).strip()
                except:
                    continue
    return None


def buscar_valor_relativo(hoja, texto_clave, delta_filas=0, delta_cols=1):
    """Busca el texto clave y devuelve el valor en una posición relativa."""
    for fila in hoja.iter_rows():
        for celda in fila:
            try:
                if celda.value and isinstance(celda.value, str) and texto_clave.lower() in celda.value.lower():
                    fila_dest = celda.row + delta_filas
                    col_dest  = celda.column + delta_cols
                    valor = hoja.cell(row=fila_dest, column=col_dest).value
                    if valor:
                        return str(valor).strip()
            except:
                continue
    return None


def obtener_hoja_portada(wb):
    """Obtiene la hoja de portada del workbook."""
    for nombre in wb.sheetnames:
        if 'portada' in nombre.lower():
            return wb[nombre]
    return None


def limpiar_fecha(valor):
    """Intenta parsear una fecha desde múltiples formatos."""
    if not valor:
        return None
    if isinstance(valor, datetime):
        return valor.date()
    formatos = ['%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%m/%d/%Y', '%Y%m%d']
    valor_str = str(valor).strip()
    for fmt in formatos:
        try:
            return datetime.strptime(valor_str, fmt).date()
        except:
            continue
    return None


def extraer_anio(fecha_emision, ruta_archivo):
    """Extrae el año de la fecha de emisión, o del path del archivo como fallback."""
    if fecha_emision:
        return fecha_emision.year
    # Fallback: buscar año en la ruta del archivo (ej: .../2017/...)
    for parte in ruta_archivo.replace('\\', '/').split('/'):
        if parte.isdigit() and 2000 <= int(parte) <= 2100:
            return int(parte)
    return datetime.now().year


def extraer_datos_certificado(ruta_archivo):
    """
    Abre un archivo Excel y extrae todos los campos del certificado.
    Retorna un dict con los datos o None si no se puede procesar.
    """
    try:
        wb = load_workbook(ruta_archivo, data_only=True)
        ws = obtener_hoja_portada(wb)

        if not ws:
            return None, "No se encontró hoja 'portada'"

        fecha_emision    = limpiar_fecha(buscar_valor_en_fila(ws, "Fecha de emisión"))
        fecha_recepcion  = limpiar_fecha(buscar_valor_en_fila(ws, "Fecha de recepción"))
        fecha_calibracion = limpiar_fecha(buscar_valor_en_fila(ws, "Fecha de calibración"))

        datos = {
            "numero_informe":          buscar_valor_relativo(ws, "Hoja 1 de 2", delta_filas=1, delta_cols=0),
            "nombre_cliente":          buscar_valor_en_fila(ws, "Nombre del cliente"),
            "direccion":               buscar_valor_en_fila(ws, "Dirección"),
            "atencion_a":              buscar_valor_en_fila(ws, "Atención a"),
            "descripcion_instrumento": buscar_valor_en_fila(ws, "Descripción del instrumento"),
            "alcance":                 buscar_valor_en_fila(ws, "Alcance"),
            "numero_serie":            buscar_valor_relativo(ws, "No. de serie", delta_filas=0, delta_cols=2),
            "identificacion":          buscar_valor_en_fila(ws, "Identificación"),
            "modelo":                  buscar_valor_en_fila(ws, "Modelo"),
            "marca":                   buscar_valor_en_fila(ws, "Marca"),
            "magnitud_evaluada":       buscar_valor_en_fila(ws, "Magnitud evaluada"),
            "resultado_calibracion":   buscar_valor_en_fila(ws, "Resultado de calibración"),
            "incertidumbre":           buscar_valor_en_fila(ws, "Incertidumbre"),
            "temperatura":             buscar_valor_en_fila(ws, "Temperatura"),
            "humedad_relativa":        buscar_valor_en_fila(ws, "Humedad"),
            "metodo_utilizado":        buscar_valor_en_fila(ws, "Método utilizado"),
            "lugar_calibracion":       buscar_valor_en_fila(ws, "Lugar de la calibración"),
            "calibrado_por":           buscar_valor_en_fila(ws, "Calibró"),
            "aprobado_por":            buscar_valor_en_fila(ws, "Aprobó"),
            "fecha_recepcion":         fecha_recepcion,
            "fecha_calibracion":       fecha_calibracion,
            "fecha_emision":           fecha_emision,
            "ruta_archivo_origen":     ruta_archivo,
            "importado_por":           "script_python",
        }

        datos["anio_emision"] = extraer_anio(fecha_emision, ruta_archivo)

        return datos, None

    except Exception as e:
        return None, str(e)


# ============================================================
# Importación principal
# ============================================================
def importar(carpetas, conn):
    cursor = conn.cursor()

    total       = 0
    exitosos    = 0
    fallidos    = 0
    errores     = []
    lote        = []

    def insertar_lote(lote):
        if not lote:
            return
        columnas = [
            'numero_informe', 'anio_emision', 'nombre_cliente', 'direccion',
            'atencion_a', 'descripcion_instrumento', 'alcance', 'numero_serie',
            'identificacion', 'modelo', 'marca', 'magnitud_evaluada',
            'resultado_calibracion', 'incertidumbre', 'temperatura', 'humedad_relativa',
            'fecha_recepcion', 'fecha_calibracion', 'fecha_emision',
            'metodo_utilizado', 'lugar_calibracion', 'calibrado_por', 'aprobado_por',
            'ruta_archivo_origen', 'importado_por'
        ]
        valores = [tuple(d.get(c) for c in columnas) for d in lote]
        sql = f"""
            INSERT INTO certificados ({', '.join(columnas)})
            VALUES %s
            ON CONFLICT DO NOTHING
        """
        execute_values(cursor, sql, valores)
        conn.commit()

    for carpeta_raiz in carpetas:
        if not os.path.exists(carpeta_raiz):
            log.warning(f"Carpeta no encontrada: {carpeta_raiz}")
            continue

        log.info(f"\nExplorando: {carpeta_raiz}")

        # Recolectar todos los archivos primero para la barra de progreso
        todos_archivos = []
        for carpeta_actual, _, archivos in os.walk(carpeta_raiz):
            validos = [
                os.path.join(carpeta_actual, a)
                for a in archivos
                if a.endswith(('.xlsx', '.xls')) and not a.startswith('~$')
            ]
            todos_archivos.extend(validos)

        log.info(f"Total archivos encontrados: {len(todos_archivos)}")

        for ruta_archivo in tqdm(todos_archivos, desc="Importando", unit="archivo"):
            total += 1
            datos, error = extraer_datos_certificado(ruta_archivo)

            if datos:
                lote.append(datos)
                exitosos += 1
                if len(lote) >= CHUNK_SIZE:
                    insertar_lote(lote)
                    lote.clear()
                    log.info(f"  Lote insertado | Total hasta ahora: {exitosos}")
            else:
                fallidos += 1
                errores.append({"archivo": ruta_archivo, "error": error})
                log.warning(f"  Fallido: {os.path.basename(ruta_archivo)} — {error}")

    # Insertar lote final
    insertar_lote(lote)
    cursor.close()

    return total, exitosos, fallidos, errores


# ============================================================
# Entry point
# ============================================================
if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='Importar certificados SIDEC a PostgreSQL')
    parser.add_argument('--carpeta', type=str, help='Carpeta específica a procesar')
    args = parser.parse_args()

    carpetas = [args.carpeta] if args.carpeta else CARPETAS_DEFAULT

    log.info("=" * 60)
    log.info("  SIDEC — Importación masiva de certificados")
    log.info(f"  Carpetas: {carpetas}")
    log.info(f"  Base de datos: {DB_CONFIG['dbname']} en {DB_CONFIG['host']}")
    log.info("=" * 60)

    try:
        conn = psycopg2.connect(**DB_CONFIG)
        log.info("✅ Conectado a PostgreSQL")
    except Exception as e:
        log.error(f"❌ No se pudo conectar a la base de datos: {e}")
        sys.exit(1)

    inicio = datetime.now()
    total, exitosos, fallidos, errores = importar(carpetas, conn)
    conn.close()

    duracion = datetime.now() - inicio

    # Guardar log de errores
    if errores:
        with open(f'errores_{datetime.now().strftime("%Y%m%d_%H%M%S")}.json', 'w', encoding='utf-8') as f:
            json.dump(errores, f, ensure_ascii=False, indent=2, default=str)

    print("\n" + "=" * 60)
    print("  RESUMEN FINAL")
    print("=" * 60)
    print(f"  Total archivos procesados : {total}")
    print(f"  ✅ Exitosos               : {exitosos}")
    print(f"  ❌ Fallidos               : {fallidos}")
    print(f"  ⏱  Tiempo total           : {duracion}")
    print("=" * 60)
